# pip install openpyxl
# pip install xlrd
# pip install pathlib
from tkinter import *
from tkinter.ttk import Combobox
from tkinter import messagebox
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib

root = Tk()
root.title("Data Entry")
root.geometry('700x400+300+200')
root.resizable(False,False)
root.configure(bg="#326273")

#Files
file=pathlib.Path('BackendData.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Full Name"
    sheet['B1']="PhoneNumber"
    sheet['C1']="Age"
    sheet['D1']="Gender"
    sheet['E1']="Address"

    file.save('BackendData.xlsx')


#functions

def submit():
    name=nameValue.get()
    contact = contactValue.get()
    age=AgeValue.get()
    gender=gender_combobox.get()
    address=addressEntry.get(1.0,END)

    file=openpyxl.load_workbook('BackendData.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row,value=contact)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row,value=address)

    file.save(r'BackendData.xlsx')

    messagebox.showinfo('info','detail added')

    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0,END)



def fetch_data():
    try:
        file = openpyxl.load_workbook('BackendData.xlsx')
        sheet = file.active

        # Assuming you want to fetch the last row of data (most recent entry)
        last_row = sheet.max_row

        # Retrieve data from Excel
        name_value = sheet.cell(row=last_row, column=1).value
        contact_value = sheet.cell(row=last_row, column=2).value
        age_value = sheet.cell(row=last_row, column=3).value
        gender_value = sheet.cell(row=last_row, column=4).value
        address_value = sheet.cell(row=last_row, column=5).value

        # Populate the form with fetched data
        nameValue.set(name_value)
        contactValue.set(contact_value)
        AgeValue.set(age_value)
        gender_combobox.set(gender_value)
        addressEntry.delete(1.0, END)
        addressEntry.insert(INSERT, address_value)

        messagebox.showinfo('Info', 'Data fetched successfully.')
    except Exception as e:
        messagebox.showerror('Error', f'Error fetching data: {str(e)}')






def clear():
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0,END)






#heading
Label(root,text="Please fill out this Entry form:",font="arial 13",bg="#326273",fg="#fff").place(x=20,y=20)


#Label
Label(root,text="Name",font=23,bg="#326273",fg="#fff").place(x=50,y=100)
Label(root,text="Contact No",font=23,bg="#326273",fg="#fff").place(x=40,y=150)
Label(root,text="Age",font=23,bg="#326273",fg="#fff").place(x=50,y=200)
Label(root,text="Gender",font=23,bg="#326273",fg="#fff").place(x=370,y=200)
Label(root,text="Address",font=23,bg="#326273",fg="#fff").place(x=50,y=250)

#Entry
nameValue = StringVar()
contactValue = StringVar()
AgeValue = StringVar()

nameEntry = Entry(root,textvariable=nameValue,width=45,font=20).place(x=150,y=100)
contactEntry = Entry(root,textvariable=contactValue,width=35,bd=2,font=20).place(x=150,y=150)
AgeEntry = Entry(root,textvariable=AgeValue,width=10,bd=2,font=20).place(x=150,y=200)

#gender
gender_combobox = Combobox(root,values=['Male','Female','Other'],font="arial 14",state="r",width=14)
gender_combobox.place(x=440,y=200)
gender_combobox.set('Male')



#Address
addressEntry = Text(root,width=50,height=5,bd=2)
addressEntry.place(x=150,y=250)



#Button
Button(root,text="Submit",bg="#326273",fg="white",width=15,height=2,command=submit).place(x=160,y=350)
Button(root,text="Clear",bg="#326273",fg="white",width=15,height=2,command=clear).place(x=300,y=350)
Button(root,text="Exit",bg="#326273",fg="white",width=15,height=2,command=lambda:root.destroy()).place(x=440,y=350)
# Button to fetch data
fetch_button = Button(root, text="Fetch Data", bg="#326273", fg="white", width=15, height=2, command=fetch_data)
fetch_button.place(x=570, y=350)



root.mainloop()